from __future__ import annotations

import csv
import subprocess
import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "exposure_review_packet_synthetic.xlsx"
HISTORICAL = ROOT / "historical_exposure_snapshots_2026-Q2.csv"
PDF = ROOT / "investment_committee_memo_synthetic.pdf"


def _run(script_name):
    subprocess.run([sys.executable, str(ROOT / "scripts" / script_name)], check=True, cwd=ROOT)


def _ensure_artifacts():
    if not WORKBOOK.exists():
        _run("generate_exposure_workbook.py")
    if not HISTORICAL.exists():
        _run("generate_historical_snapshots.py")
    if not PDF.exists():
        _run("generate_market_context_memo_pdf.py")


def _sheet_rows(ws):
    headers = [cell.value for cell in ws[1]]
    return [dict(zip(headers, values)) for values in ws.iter_rows(min_row=2, values_only=True)]


def test_exports_and_intended_demo_behavior():
    _ensure_artifacts()

    assert WORKBOOK.exists()
    assert HISTORICAL.exists()
    assert PDF.exists()
    assert PDF.stat().st_size > 1000

    wb = load_workbook(WORKBOOK, data_only=True)
    expected_sheets = {
        "Overview",
        "Positions_Raw",
        "Position_View",
        "Review_Thresholds",
        "Source_Links_Raw",
        "Data_Dictionary",
        "Pipeline_QA",
    }
    assert expected_sheets.issubset(set(wb.sheetnames))

    positions = _sheet_rows(wb["Positions_Raw"])
    assert len(positions) == 30

    required_columns = {
        "as_of_date",
        "portfolio_id",
        "ticker",
        "instrument_name",
        "asset_class",
        "sector",
        "issuer",
        "canonical_issuer_id",
        "quantity",
        "price",
        "market_value",
        "gross_exposure",
        "issuer_limit",
        "currency",
        "source_ids",
        "is_synthetic",
    }
    assert required_columns.issubset(set(positions[0]))

    source_rows = _sheet_rows(wb["Source_Links_Raw"])
    source_ids = {row["source_id"] for row in source_rows}
    for row in positions:
        assert row["is_synthetic"] is True
        for source_id in str(row["source_ids"]).split("|"):
            assert source_id in source_ids

    agrd = next(row for row in positions if row["ticker"] == "AGRD")
    agrd_cnv = next(row for row in positions if row["ticker"] == "AGRD-CNV")
    assert agrd["canonical_issuer_id"] == "ISS-104"
    assert agrd_cnv["canonical_issuer_id"] == "ISS-104"

    exposure_by_issuer = defaultdict(float)
    limit_by_issuer = {}
    for row in positions:
        exposure_by_issuer[row["canonical_issuer_id"]] += float(row["gross_exposure"])
        limit_by_issuer[row["canonical_issuer_id"]] = float(row["issuer_limit"])

    assert round(exposure_by_issuer["ISS-104"], 1) == 15.2
    assert exposure_by_issuer["ISS-104"] > limit_by_issuer["ISS-104"]
    other_breaches = [
        issuer_id
        for issuer_id, exposure in exposure_by_issuer.items()
        if issuer_id != "ISS-104" and exposure > limit_by_issuer[issuer_id]
    ]
    assert other_breaches == []

    with open(HISTORICAL, newline="", encoding="utf-8") as handle:
        historical_rows = list(csv.DictReader(handle))
    assert len(historical_rows) == 210
    assert {row["ticker"] for row in historical_rows} == {row["ticker"] for row in positions}


if __name__ == "__main__":
    test_exports_and_intended_demo_behavior()
    print("PASS: export artifacts and intended demo behavior validated")
